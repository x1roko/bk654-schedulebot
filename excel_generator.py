import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict

# Стили для Excel
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
CELL_ALIGNMENT = Alignment(horizontal='center', vertical='center')
WORKING_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DAY_OFF_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def generate_week_schedule_excel(schedule_data: List[Dict], week_start_date: datetime.date) -> str:
    """Генерирует Excel файл с расписанием на неделю"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    # Заголовки
    days_ru = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    days_en = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    headers = ['Фамилия Имя'] + days_ru
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CELL_ALIGNMENT
    
    # Записываем данные
    for row_num, entry in enumerate(schedule_data, 2):
        name = f"{entry['last_name']} {entry['first_name']}"
        ws.cell(row=row_num, column=1, value=name)
        
        for col_num, day in enumerate(days_en, 2):
            value = entry.get(day, 'выходной')
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = CELL_ALIGNMENT
            
            if value.lower() == 'выходной':
                cell.fill = DAY_OFF_FILL
            else:
                cell.fill = WORKING_FILL
    
    # Настраиваем ширину столбцов
    for col in ws.columns:
        column = get_column_letter(col[0].column)
        ws.column_dimensions[column].width = 15 if column != 'A' else 20
    
    # Сохраняем файл
    filename = f"schedule_{week_start_date.strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    return filename

def generate_day_schedule_excel(schedule_entries: List[Dict], day_name: str) -> str:
    """Генерирует Excel файл с расписанием на конкретный день"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Расписание на {day_name}"
    
    # Заголовки
    hours = [f"{hour:02d}:00" for hour in range(8, 23)]
    headers = ['Фамилия Имя'] + hours
    
    # Записываем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CELL_ALIGNMENT
    
    # Записываем данные
    for row_num, entry in enumerate(schedule_entries, 2):
        name = f"{entry['last_name']} {entry['first_name']}"
        ws.cell(row=row_num, column=1, value=name)
        
        time_range = entry.get(day_name, 'выходной')
        
        if time_range.lower() == 'выходной':
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=len(headers))
            cell = ws.cell(row=row_num, column=2, value="Выходной")
            cell.alignment = CELL_ALIGNMENT
            cell.fill = DAY_OFF_FILL
            continue
            
        try:
            start_time, end_time = time_range.split('-')
            start_hour = int(start_time.split(':')[0]) if ':' in start_time else int(start_time)
            end_hour = int(end_time.split(':')[0]) if ':' in end_time else int(end_time)
            
            for col_num, hour in enumerate(range(8, 23), 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = CELL_ALIGNMENT
                
                if start_hour <= hour < end_hour:
                    cell.value = "✓"
                    cell.fill = WORKING_FILL
                else:
                    cell.value = "✗"
                    cell.fill = DAY_OFF_FILL
        except ValueError:
            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=len(headers))
            cell = ws.cell(row=row_num, column=2, value="Некорректный формат")
            cell.alignment = CELL_ALIGNMENT
    
    # Настраиваем ширину столбцов
    for col in ws.columns:
        column = get_column_letter(col[0].column)
        ws.column_dimensions[column].width = 5 if column != 'A' else 20
    
    # Сохраняем файл
    filename = f"schedule_{day_name}.xlsx"
    wb.save(filename)
    return filename
